import csv
import datetime
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .models import Income
from .serializers import IncomeSerializer
from .filters import IncomeFilter


class IncomeViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for Income records.
    Extra actions:
      GET /api/incomes/summary/        — monthly + yearly totals, breakdown by source
      GET /api/incomes/monthly-trend/  — last 12 months income by month
      GET /api/incomes/yearly-summary/ — income totals grouped by year
    """
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = IncomeFilter
    search_fields = ['title', 'description']
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date']

    def get_queryset(self):
        qs = Income.objects.filter(user=self.request.user)

        # Date range filters
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')

        if start_date:
            qs = qs.filter(date__gte=start_date)
        if end_date:
            qs = qs.filter(date__lte=end_date)
        if year:
            qs = qs.filter(date__year=year)
        if month:
            qs = qs.filter(date__month=month)

        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Returns monthly income, yearly income, and breakdown by source."""
        user = request.user
        today = timezone.now().date()
        month_start = today.replace(day=1)
        year_start = today.replace(month=1, day=1)

        all_incomes = Income.objects.filter(user=user)

        # Monthly totals
        month_qs = all_incomes.filter(date__gte=month_start, date__lte=today)
        monthly_total = float(month_qs.aggregate(total=Sum('amount'))['total'] or 0)
        monthly_count = month_qs.count()

        # Yearly totals
        year_qs = all_incomes.filter(date__gte=year_start, date__lte=today)
        yearly_total = float(year_qs.aggregate(total=Sum('amount'))['total'] or 0)
        yearly_count = year_qs.count()

        # All-time total
        all_time_total = float(all_incomes.aggregate(total=Sum('amount'))['total'] or 0)

        # Breakdown by source (current year)
        SOURCE_META = {
            'salary':      {'icon': '💼', 'color': '#6366f1'},
            'business':    {'icon': '🏢', 'color': '#10b981'},
            'investment':  {'icon': '📈', 'color': '#f59e0b'},
            'freelancing': {'icon': '💻', 'color': '#06b6d4'},
            'other':       {'icon': '💰', 'color': '#ec4899'},
        }

        source_breakdown = (
            year_qs
            .values('source')
            .annotate(total=Sum('amount'))
            .order_by('-total')
        )

        breakdown_data = []
        for item in source_breakdown:
            src = item['source']
            meta = SOURCE_META.get(src, {'icon': '💰', 'color': '#6b7280'})
            total = float(item['total'])
            breakdown_data.append({
                'source': src,
                'source_display': dict(Income.SOURCE_CHOICES).get(src, src.title()),
                'icon': meta['icon'],
                'color': meta['color'],
                'total': total,
                'percentage': round((total / yearly_total) * 100, 1) if yearly_total > 0 else 0,
            })

        return Response({
            'monthly_total': monthly_total,
            'monthly_count': monthly_count,
            'yearly_total': yearly_total,
            'yearly_count': yearly_count,
            'all_time_total': all_time_total,
            'source_breakdown': breakdown_data,
        })

    @action(detail=False, methods=['get'], url_path='monthly-trend')
    def monthly_trend(self, request):
        """Returns income totals for the last 12 months."""
        user = request.user
        twelve_months_ago = timezone.now().date() - datetime.timedelta(days=365)

        trends = (
            Income.objects.filter(user=user, date__gte=twelve_months_ago)
            .annotate(month=TruncMonth('date'))
            .values('month')
            .annotate(total=Sum('amount'))
            .order_by('month')
        )

        data = [
            {
                'month': item['month'].strftime('%Y-%m') if item['month'] else '',
                'month_display': item['month'].strftime('%b %Y') if item['month'] else '',
                'total': float(item['total']),
            }
            for item in trends
        ]
        return Response(data)

    @action(detail=False, methods=['get'], url_path='yearly-summary')
    def yearly_summary(self, request):
        """Returns income totals grouped by year."""
        user = request.user

        from django.db.models.functions import ExtractYear
        year_data = (
            Income.objects.filter(user=user)
            .annotate(year=ExtractYear('date'))
            .values('year')
            .annotate(total=Sum('amount'))
            .order_by('year')
        )

        data = [
            {'year': item['year'], 'total': float(item['total'])}
            for item in year_data
        ]
        return Response(data)

    # ─── Export Actions ─────────────────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        incomes = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="incomes_{datetime.date.today()}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Title', 'Source', 'Amount', 'Recurring', 'Recurrence Period', 'Description'])

        for inc in incomes:
            writer.writerow([
                inc.date.strftime('%Y-%m-%d') if inc.date else '',
                inc.title,
                inc.get_source_display(),
                float(inc.amount),
                'Yes' if inc.is_recurring else 'No',
                inc.get_recurrence_period_display() if inc.recurrence_period else '',
                inc.description or '',
            ])

        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        incomes = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Income Summary'

        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0F6B3D', end_color='0F6B3D', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Date', 'Title', 'Source', 'Amount', 'Recurring', 'Recurrence Period', 'Description']
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 35

        total_sum = 0
        for inc in incomes:
            ws.append([
                inc.date.strftime('%Y-%m-%d') if inc.date else '',
                inc.title,
                inc.get_source_display(),
                float(inc.amount),
                'Yes' if inc.is_recurring else 'No',
                inc.get_recurrence_period_display() if inc.recurrence_period else '',
                inc.description or '',
            ])
            total_sum += inc.amount

        row_count = ws.max_row
        ws.cell(row=row_count + 2, column=3, value='Total Income:').font = Font(bold=True)
        total_cell = ws.cell(row=row_count + 2, column=4, value=float(total_sum))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '$#,##0.00'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="incomes_{datetime.date.today()}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        incomes = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="incomes_{datetime.date.today()}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=colors.HexColor('#0F6B3D'),
            alignment=1,
            spaceAfter=20
        )
        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#555555'),
            spaceAfter=15
        )

        elements.append(Paragraph('Personal Income Report', title_style))
        elements.append(Paragraph(
            f"Generated on: {datetime.date.today().strftime('%B %d, %Y')} | Total records: {incomes.count()}",
            meta_style
        ))
        elements.append(Spacer(1, 10))

        data = [['Date', 'Title', 'Source', 'Recurring', 'Amount']]
        total_sum = 0
        for inc in incomes:
            data.append([
                inc.date.strftime('%Y-%m-%d'),
                Paragraph(inc.title or '', styles['Normal']),
                inc.get_source_display(),
                'Yes' if inc.is_recurring else 'No',
                f'${inc.amount:,.2f}',
            ])
            total_sum += inc.amount

        data.append(['', '', '', 'Total Income:', f'${total_sum:,.2f}'])

        t = Table(data, colWidths=[70, 160, 100, 70, 90])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F6B3D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F2F9F6')),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#0F6B3D')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
        ]))

        elements.append(t)
        doc.build(elements)
        return response
