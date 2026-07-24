import csv
import datetime
from django.http import HttpResponse, StreamingHttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .models import Expense, RecurringExpense
from .serializers import ExpenseSerializer, RecurringExpenseSerializer
from .filters import ExpenseFilter
from .tasks import process_receipt_ocr

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    filterset_class = ExpenseFilter
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['title', 'description', 'category__name', 'payment_method']
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date']

    def get_queryset(self):
        from django.db.models import Q
        return self.queryset.filter(
            Q(user=self.request.user) | Q(group__members__user=self.request.user)
        ).distinct()

    def perform_create(self, serializer):
        # Set user as current user
        has_image = 'receipt_image' in self.request.FILES
        ocr_status = 'pending' if has_image else 'none'
        expense = serializer.save(user=self.request.user, ocr_status=ocr_status)
        
        # Trigger OCR task if receipt image is present
        if has_image:
            process_receipt_ocr.delay(expense.id)

    def perform_update(self, serializer):
        expense = self.get_object()
        old_image = expense.receipt_image
        new_expense = serializer.save()
        
        # If receipt image is added or changed, run OCR
        if new_expense.receipt_image and new_expense.receipt_image != old_image:
            new_expense.ocr_status = 'pending'
            new_expense.save()
            process_receipt_ocr.delay(new_expense.id)

    # ─── Receipt OCR Scan (pre-fill) ─────────────────────────────────
    @action(detail=False, methods=['post'], url_path='scan-receipt')
    def scan_receipt(self, request):
        """
        POST /api/expenses/scan-receipt/
        Body: multipart/form-data, field 'image' (JPEG/PNG)

        Runs pytesseract on the uploaded image and returns extracted fields:
          { raw_text, extracted: { title, amount, date, category_id, category_name,
            payment_method, description }, confidence }
        The frontend uses this to auto-fill the expense form.
        """
        import re
        import io
        import tempfile
        from decimal import Decimal, InvalidOperation
        from PIL import Image, ImageFilter, ImageOps
        import pytesseract as pyt
        from django.conf import settings as djsettings
        from categories.models import Category

        pyt.pytesseract.tesseract_cmd = getattr(djsettings, 'TESSERACT_CMD', 'tesseract')

        image_file = request.FILES.get('image')
        if not image_file:
            return Response(
                {'detail': 'No image uploaded. Send the receipt as field "image".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowed_ext = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp')
        if not image_file.name.lower().endswith(allowed_ext):
            return Response(
                {'detail': f'Unsupported format. Accepted: {", ".join(allowed_ext)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            img = Image.open(image_file)

            # ── Image pre-processing for better OCR accuracy ──────────────
            img = ImageOps.exif_transpose(img) or img          # fix rotation
            if img.mode != 'L':
                img = img.convert('L')                          # grayscale
            img = img.filter(ImageFilter.SHARPEN)               # sharpen
            img = ImageOps.autocontrast(img, cutoff=2)          # boost contrast

            # ── Run OCR ───────────────────────────────────────────────────
            raw_text = pyt.image_to_string(img, config='--psm 6')
        except Exception as exc:
            return Response(
                {'detail': f'OCR processing failed: {str(exc)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        lines      = [ln.strip() for ln in raw_text.split('\n') if ln.strip()]
        confidence = 0
        extracted  = {
            'title': '',
            'amount': '',
            'date': '',
            'category_id': None,
            'category_name': '',
            'payment_method': '',
            'description': '',
        }

        # ── 1. Extract store / merchant name (first non-trivial line) ──
        for ln in lines[:5]:
            # Skip lines that look purely numeric or very short
            cleaned = re.sub(r'[^a-zA-Z\s]', '', ln).strip()
            if len(cleaned) >= 3:
                extracted['title'] = ln[:80]
                confidence += 1
                break

        # ── 2. Extract total amount ───────────────────────────────────
        amount_patterns = [
            re.compile(r'(?:grand\s*total|total\s*(?:due|amt|amount)?|amount\s*due|balance\s*due|net\s*total)[\s:]*[\$£€₹]?\s*(\d{1,7}[.,]\d{2})', re.IGNORECASE),
            re.compile(r'total[\s:]*[\$£€₹]?\s*(\d{1,7}[.,]\d{2})', re.IGNORECASE),
            re.compile(r'[\$£€₹]\s*(\d{1,7}[.,]\d{2})'),
            re.compile(r'(\d{1,7}\.\d{2})'),
        ]
        best_amount = None
        for pat in amount_patterns:
            for ln in reversed(lines):  # totals are usually at the bottom
                m = pat.search(ln)
                if m:
                    try:
                        val = Decimal(m.group(1).replace(',', '.'))
                        if val > 0:
                            best_amount = val
                            break
                    except InvalidOperation:
                        pass
            if best_amount is not None:
                break

        if best_amount is not None:
            extracted['amount'] = str(best_amount)
            confidence += 1

        # ── 3. Extract date ───────────────────────────────────────────
        date_patterns = [
            (re.compile(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})'),        '%Y-%m-%d'),
            (re.compile(r'(\d{1,2}[-/]\d{1,2}[-/]\d{4})'),        None),       # ambiguous
            (re.compile(r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s,]+\d{4})', re.IGNORECASE), '%d %b %Y'),
            (re.compile(r'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2}[\s,]+\d{4})', re.IGNORECASE), '%b %d %Y'),
        ]
        for pat, fmt in date_patterns:
            for ln in lines:
                dm = pat.search(ln)
                if dm:
                    raw_date = dm.group(1).replace(',', ' ').strip()
                    if fmt:
                        try:
                            parsed = datetime.datetime.strptime(raw_date, fmt).date()
                            extracted['date'] = str(parsed)
                            confidence += 1
                            break
                        except ValueError:
                            pass
                    else:
                        # Try DD/MM/YYYY then MM/DD/YYYY
                        for try_fmt in ('%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y'):
                            try:
                                parsed = datetime.datetime.strptime(raw_date, try_fmt).date()
                                extracted['date'] = str(parsed)
                                confidence += 1
                                break
                            except ValueError:
                                continue
            if extracted['date']:
                break

        # ── 4. Detect payment method ──────────────────────────────────
        text_lower = raw_text.lower()
        if any(k in text_lower for k in ('visa', 'mastercard', 'credit', 'amex')):
            extracted['payment_method'] = 'credit_card'
            confidence += 1
        elif any(k in text_lower for k in ('debit', 'maestro', 'atm')):
            extracted['payment_method'] = 'debit_card'
            confidence += 1
        elif any(k in text_lower for k in ('transfer', 'wire', 'neft', 'upi', 'ach')):
            extracted['payment_method'] = 'bank_transfer'
            confidence += 1
        elif 'cash' in text_lower:
            extracted['payment_method'] = 'cash'
            confidence += 1

        # ── 5. Auto-match category ────────────────────────────────────
        CATEGORY_KEYWORDS = {
            'food':          ['restaurant', 'cafe', 'coffee', 'pizza', 'burger', 'dine', 'food', 'eat', 'kitchen', 'bakery', 'grill', 'sushi', 'taco', 'starbucks', 'mcdonald', 'subway'],
            'groceries':     ['grocery', 'supermarket', 'market', 'walmart', 'costco', 'target', 'aldi', 'lidl', 'whole foods', 'kroger', 'safeway', 'produce'],
            'transport':     ['uber', 'lyft', 'taxi', 'fuel', 'gas', 'petrol', 'diesel', 'parking', 'transit', 'metro', 'bus', 'toll', 'shell', 'bp', 'chevron'],
            'entertainment': ['cinema', 'movie', 'theater', 'netflix', 'spotify', 'game', 'concert', 'ticket', 'amusement', 'bowling'],
            'shopping':      ['amazon', 'ebay', 'mall', 'store', 'shop', 'retail', 'clothing', 'fashion', 'nike', 'adidas', 'zara'],
            'utilities':     ['electric', 'water', 'gas bill', 'internet', 'phone', 'mobile', 'wifi', 'cable', 'utility', 'verizon', 'at&t', 'comcast'],
            'health':        ['pharmacy', 'hospital', 'doctor', 'clinic', 'medical', 'health', 'dental', 'vision', 'cvs', 'walgreens', 'prescription'],
            'housing':       ['rent', 'mortgage', 'lease', 'housing', 'apartment', 'maintenance', 'repair'],
            'education':     ['tuition', 'school', 'university', 'college', 'course', 'book', 'udemy', 'coursera'],
            'travel':        ['hotel', 'flight', 'airline', 'airbnb', 'booking', 'resort', 'travel', 'trip', 'luggage'],
        }

        user_categories = {
            c.name.strip().lower(): c
            for c in Category.objects.filter(user=request.user)
        }

        matched_cat = None
        # First try to match category keywords against OCR text
        for cat_key, keywords in CATEGORY_KEYWORDS.items():
            for kw in keywords:
                if kw in text_lower:
                    # Find user's category that matches this keyword group
                    for ucat_name, ucat_obj in user_categories.items():
                        if cat_key in ucat_name or ucat_name in cat_key:
                            matched_cat = ucat_obj
                            break
                    # Also try partial match on any user category name
                    if not matched_cat:
                        for ucat_name, ucat_obj in user_categories.items():
                            if kw in ucat_name or ucat_name in kw:
                                matched_cat = ucat_obj
                                break
                    if matched_cat:
                        break
            if matched_cat:
                break

        if matched_cat:
            extracted['category_id']   = matched_cat.id
            extracted['category_name'] = matched_cat.name
            confidence += 1

        # ── 6. Build description from receipt text ────────────────────
        # Take up to 5 item-like lines (skip the store name line)
        item_lines = []
        for ln in lines[1:]:
            if len(ln) > 5 and not any(skip in ln.lower() for skip in ('total', 'subtotal', 'tax', 'change', 'visa', 'mastercard', 'debit', 'credit', 'card', 'thank')):
                item_lines.append(ln)
            if len(item_lines) >= 5:
                break
        if item_lines:
            extracted['description'] = '\n'.join(item_lines)

        # ── Confidence score (0-5 → percentage) ──────────────────────
        max_fields   = 5  # title, amount, date, payment_method, category
        conf_percent = round((confidence / max_fields) * 100)

        return Response({
            'raw_text':   raw_text,
            'extracted':  extracted,
            'confidence': conf_percent,
        }, status=status.HTTP_200_OK)


    def export_excel(self, request):
        expenses = self.filter_queryset(self.get_queryset())
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses Summary"
        
        # Header formatting
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        headers = ["Date", "Title", "Category", "Payment Method", "Amount", "Description", "Notes", "OCR Text Extraction"]
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 40

        total_sum = 0
        for exp in expenses:
            ws.append([
                exp.date.strftime("%Y-%m-%d") if exp.date else "",
                exp.title,
                exp.category.name if exp.category else "Uncategorized",
                exp.get_payment_method_display(),
                float(exp.amount),
                exp.description or "",
                exp.notes or "",
                exp.ocr_text or ""
            ])
            total_sum += exp.amount

        # Add total row
        row_count = ws.max_row
        ws.cell(row=row_count + 2, column=4, value="Total Summary:").font = Font(bold=True)
        total_cell = ws.cell(row=row_count + 2, column=5, value=float(total_sum))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '$#,##0.00'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="expenses_{datetime.date.today()}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        expenses = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="expenses_{datetime.date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []
        
        # PDF Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=colors.HexColor('#1F4E78'),
            alignment=1, # Center
            spaceAfter=20
        )
        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#555555'),
            spaceAfter=15
        )
        
        # Title and Header
        elements.append(Paragraph("Personal Expense Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.date.today().strftime('%B %d, %Y')} | Total records: {expenses.count()}", meta_style))
        elements.append(Spacer(1, 10))
        
        # Table data
        data = [["Date", "Title", "Category", "Method", "Amount"]]
        total_sum = 0
        for exp in expenses:
            data.append([
                exp.date.strftime("%Y-%m-%d"),
                Paragraph(exp.title or "", styles['Normal']),
                exp.category.name if exp.category else "Uncategorized",
                exp.get_payment_method_display(),
                f"${exp.amount:,.2f}"
            ])
            total_sum += exp.amount
            
        data.append(["", "", "", "Total Amount:", f"${total_sum:,.2f}"])
        
        t = Table(data, colWidths=[70, 150, 100, 100, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-2), colors.HexColor('#F2F2F2')),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CCCCCC')),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1F4E78')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('TOPPADDING', (0,-1), (-1,-1), 8),
        ]))
        
        elements.append(t)
        doc.build(elements)
        return response

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        expenses = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="expenses_{datetime.date.today()}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Title', 'Category', 'Payment Method', 'Amount', 'Description', 'Notes', 'OCR Status', 'OCR Text'])

        for exp in expenses:
            writer.writerow([
                exp.date.strftime('%Y-%m-%d') if exp.date else '',
                exp.title,
                exp.category.name if exp.category else 'Uncategorized',
                exp.get_payment_method_display(),
                float(exp.amount),
                exp.description or '',
                exp.notes or '',
                exp.get_ocr_status_display(),
                exp.ocr_text or '',
            ])

        return response

    # ─── CSV Import ──────────────────────────────────────────────────────
    IMPORT_REQUIRED_COLUMNS = {'date', 'title', 'amount'}
    VALID_PAYMENT_METHODS = {
        'cash': 'cash',
        'credit card': 'credit_card',
        'credit_card': 'credit_card',
        'debit card': 'debit_card',
        'debit_card': 'debit_card',
        'bank transfer': 'bank_transfer',
        'bank_transfer': 'bank_transfer',
        'others': 'others',
        'other': 'others',
    }

    @action(detail=False, methods=['post'], url_path='import-csv')
    def import_csv(self, request):
        """
        POST /api/expenses/import-csv/
        Body: multipart/form-data with field 'file' (CSV).

        Accepted columns (case-insensitive):
          date, title, category, payment_method, amount, description, notes

        Returns a JSON import report:
          { total, imported, skipped, errors, rows: [{row, status, reason, data}] }
        """
        from decimal import Decimal, InvalidOperation
        from categories.models import Category

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'detail': 'No file uploaded. Send the CSV as field "file".'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file_obj.name.lower().endswith('.csv'):
            return Response(
                {'detail': 'Only .csv files are accepted.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Decode ────────────────────────────────────────────────────────
        try:
            raw = file_obj.read().decode('utf-8-sig')  # strip BOM if present
        except UnicodeDecodeError:
            return Response(
                {'detail': 'File encoding not supported. Please use UTF-8.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        reader = csv.DictReader(raw.splitlines())

        # Validate header row exists
        if reader.fieldnames is None:
            return Response(
                {'detail': 'CSV file is empty or has no header row.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        headers_normalized = {h.strip().lower() for h in reader.fieldnames}
        missing = self.IMPORT_REQUIRED_COLUMNS - headers_normalized
        if missing:
            return Response(
                {'detail': f'CSV is missing required column(s): {", ".join(sorted(missing))}. Required: date, title, amount.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Build category lookup for this user (case-insensitive name -> obj) ──
        user_categories = {
            c.name.strip().lower(): c
            for c in Category.objects.filter(user=request.user)
        }

        # ── Process rows ──────────────────────────────────────────────────
        report_rows = []
        imported_count = 0
        skipped_count  = 0
        error_count    = 0

        for raw_row_num, raw_row in enumerate(reader, start=2):  # row 1 is header
            # Normalize keys
            row = {k.strip().lower(): (v.strip() if v else '') for k, v in raw_row.items()}
            row_errors = []

            # ── Validate date ─────────────────────────────────────────────
            date_str = row.get('date', '')
            parsed_date = None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    parsed_date = datetime.datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed_date is None:
                row_errors.append(f'Invalid date "{date_str}". Use YYYY-MM-DD.')

            # ── Validate title ────────────────────────────────────────────
            title = row.get('title', '').strip()
            if not title:
                row_errors.append('Title is required.')

            # ── Validate amount ───────────────────────────────────────────
            amount_raw = row.get('amount', '')
            # Strip common currency symbols and commas
            for sym in ('$', '£', '€', '₹', ','):
                amount_raw = amount_raw.replace(sym, '')
            amount_raw = amount_raw.strip()
            parsed_amount = None
            try:
                parsed_amount = Decimal(amount_raw)
                if parsed_amount <= 0:
                    row_errors.append(f'Amount must be > 0 (got {amount_raw}).')
            except InvalidOperation:
                row_errors.append(f'Invalid amount "{row.get("amount", "")}".')

            # ── Resolve payment method ────────────────────────────────────
            pm_raw = row.get('payment_method', row.get('payment method', 'cash')).strip().lower()
            payment_method = self.VALID_PAYMENT_METHODS.get(pm_raw, 'cash')

            # ── Resolve category ──────────────────────────────────────────
            cat_name = row.get('category', '').strip().lower()
            category_obj = user_categories.get(cat_name) if cat_name else None

            # ── Optional fields ───────────────────────────────────────────
            description = row.get('description', '')
            notes       = row.get('notes', '')

            # ── Validation failed → error ─────────────────────────────────
            if row_errors:
                error_count += 1
                report_rows.append({
                    'row':    raw_row_num,
                    'status': 'error',
                    'reason': '; '.join(row_errors),
                    'data':   dict(raw_row),
                })
                continue

            # ── Duplicate check: same user + date + amount + title ────────
            duplicate_exists = Expense.objects.filter(
                user=request.user,
                date=parsed_date,
                amount=parsed_amount,
                title__iexact=title,
            ).exists()

            if duplicate_exists:
                skipped_count += 1
                report_rows.append({
                    'row':    raw_row_num,
                    'status': 'skipped',
                    'reason': f'Duplicate: "{title}" on {parsed_date} for {parsed_amount} already exists.',
                    'data':   dict(raw_row),
                })
                continue

            # ── Create expense ────────────────────────────────────────────
            Expense.objects.create(
                user=request.user,
                title=title,
                category=category_obj,
                payment_method=payment_method,
                amount=parsed_amount,
                date=parsed_date,
                description=description,
                notes=notes,
                ocr_status='none',
            )
            imported_count += 1
            report_rows.append({
                'row':    raw_row_num,
                'status': 'imported',
                'reason': '',
                'data': {
                    'title':          title,
                    'date':           str(parsed_date),
                    'amount':         str(parsed_amount),
                    'category':       category_obj.name if category_obj else 'Uncategorized',
                    'payment_method': payment_method,
                },
            })

        total = imported_count + skipped_count + error_count
        return Response({
            'total':    total,
            'imported': imported_count,
            'skipped':  skipped_count,
            'errors':   error_count,
            'rows':     report_rows,
        }, status=status.HTTP_200_OK)


class RecurringExpenseViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for recurring expense rules.
    Extra actions:
      POST /api/recurring-expenses/{id}/toggle/    — toggle is_active
      POST /api/recurring-expenses/{id}/trigger/   — manually generate the next expense now
    """
    queryset = RecurringExpense.objects.all()
    serializer_class = RecurringExpenseSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['title', 'description', 'category__name']
    ordering_fields = ['next_due_date', 'amount', 'created_at']
    ordering = ['next_due_date']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        data = serializer.validated_data
        # Default next_due_date to start_date
        if not data.get('next_due_date'):
            data['next_due_date'] = data.get('start_date')
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'], url_path='toggle')
    def toggle_active(self, request, pk=None):
        """Pause or resume a recurring expense."""
        recurring = self.get_object()
        recurring.is_active = not recurring.is_active
        recurring.save(update_fields=['is_active'])
        return Response(RecurringExpenseSerializer(recurring).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='trigger')
    def trigger_now(self, request, pk=None):
        """
        Manually generate the next expense from this recurring rule right now,
        regardless of next_due_date.
        """
        recurring = self.get_object()
        today = datetime.date.today()

        # Create the expense
        expense = Expense.objects.create(
            user=recurring.user,
            title=recurring.title,
            category=recurring.category,
            payment_method=recurring.payment_method,
            amount=recurring.amount,
            date=today,
            description=recurring.description or '',
            notes=f'[Auto-generated] Recurring: {recurring.get_frequency_display()} — triggered manually',
            ocr_status='none',
            recurring_expense=recurring,
        )

        # Advance next_due_date
        next_due = recurring.compute_next_due(from_date=today)
        if next_due is None:
            recurring.is_active = False
        else:
            recurring.next_due_date = next_due

        recurring.total_generated += 1
        recurring.last_generated = today
        recurring.save()

        return Response({
            'status': 'generated',
            'expense_id': expense.id,
            'next_due_date': str(recurring.next_due_date) if recurring.is_active else None,
            'recurring': RecurringExpenseSerializer(recurring).data,
        }, status=status.HTTP_200_OK)
